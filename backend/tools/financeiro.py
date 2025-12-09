import pandas as pd
from datetime import datetime

def normalizar_planilha_financeira(entrada):
    """
    Normaliza a planilha financeira conforme os requisitos especificados.
    
    Requisitos implementados:
    1) Colocar C na frente do código
    2) Segmentar primeira coluna em código e cliente
    3) Código com letra C antes
    4) Código com 8 caracteres (código base 6 dígitos + loja 2 dígitos)
    5) Aglutinar em: código, cliente, valor (Tit Vencidos Valor Corrigido)
    6) Criar coluna TIPO baseada em dias de vencimento (>365 = LONGO PRAZO, <=365 = CURTO PRAZO)
    
    Parâmetros:
    -----------
    entrada : str ou DataFrame
        Pode ser:
        - Caminho para o arquivo Excel (string)
        - DataFrame pandas já carregado
    
    Retorna:
    --------
    DataFrame pandas normalizado
    """
    
    # Verificar se entrada é um DataFrame ou caminho de arquivo
    if isinstance(entrada, pd.DataFrame):
        print("Usando DataFrame fornecido...")
        df = entrada.copy()
    elif isinstance(entrada, str):
        print("Lendo arquivo...")
        df = pd.read_excel(entrada)
    else:
        raise ValueError("Entrada deve ser um caminho de arquivo (string) ou DataFrame pandas")
    
    print(f"Total de registros lidos: {len(df)}")
    
    # 2) Segmentar primeira coluna em código e cliente
    # Formato original: "000672-01-A A DANTAS RIBEIRO"
    # Onde: 000672 = código base (6 dígitos)
    #       01 = loja (2 dígitos)
    #       A A DANTAS RIBEIRO = nome do cliente
    
    primeira_coluna = 'Codigo-Lj-Nome do Cliente'
    
    # Dividir por '-' limitando a 2 divisões (3 partes no total)
    df_split = df[primeira_coluna].str.split('-', n=2, expand=True)
    
    # 4) Código completo: código base (6 dígitos) + loja (2 dígitos) = 8 caracteres
    # Remover qualquer "-" que possa existir e garantir 8 caracteres
    codigo_base = df_split[0].str.strip()  # Ex: "000672"
    loja = df_split[1].str.strip()         # Ex: "01"
    
    # Concatenar código base + loja = 8 dígitos
    df['codigo_limpo'] = (codigo_base + loja).str[:8]
    
    # 1 e 3) Adicionar "C" na frente do código
    df['codigo'] = 'C' + df['codigo_limpo']
    
    # Cliente: terceira parte após o segundo "-"
    df['cliente'] = df_split[2].str.strip()
    
    # 5) Selecionar coluna de valor: "Tit Vencidos Valor Corrigido"
    df['valor'] = df['Tit Vencidos Valor Corrigido']
    
    # 6) Criar coluna TIPO baseado na data de vencimento
    # Usar "Vencto Real" como data de vencimento
    df['data_vencimento'] = pd.to_datetime(df['Vencto Real'], errors='coerce')
    
    # Calcular diferença em dias até hoje
    hoje = datetime.now()
    df['dias_vencidos'] = (hoje - df['data_vencimento']).dt.days
    
    # Criar DataFrame temporário com todas as informações necessárias
    df_temp = df[['codigo', 'cliente', 'valor', 'dias_vencidos']].copy()
    
    # Remover linhas onde o valor é NaN
    df_temp = df_temp[df_temp['valor'].notna()].copy()
    
    # Aglutinar por código: somar valores e pegar informações para calcular TIPO
    df_agrupado = df_temp.groupby(['codigo', 'cliente']).agg({
        'valor': 'sum',
        'dias_vencidos': 'max',     # Maior número de dias vencidos
    }).reset_index()
    
    # Calcular coluna TIPO após agrupamento
    df_agrupado['TIPO'] = df_agrupado['dias_vencidos'].apply(
        lambda x: 'LONGO PRAZO' if pd.notna(x) and x > 365 else 'CURTO PRAZO'
    )
    
    # Criar DataFrame final apenas com as colunas solicitadas
    df_final = df_agrupado[['codigo', 'cliente', 'valor', 'TIPO']].copy()
    
    return df_final


def salvar_planilha_normalizada(df, caminho_saida):
    """
    Salva o DataFrame normalizado em um arquivo Excel com formatação de moeda.
    
    Parâmetros:
    -----------
    df : DataFrame
        DataFrame normalizado
    caminho_saida : str
        Caminho para salvar o arquivo de saída
    """
    from openpyxl import load_workbook
    from openpyxl.styles import numbers
    
    # Salvar o DataFrame no Excel
    df.to_excel(caminho_saida, index=False, engine='openpyxl')
    
    # Carregar o workbook para aplicar formatação
    wb = load_workbook(caminho_saida)
    ws = wb.active
    
    # Encontrar a coluna "valor" (normalmente é a coluna C - índice 3)
    coluna_valor = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'valor':
            coluna_valor = idx
            break
    
    # Aplicar formatação de moeda brasileira na coluna valor
    if coluna_valor:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coluna_valor)
            cell.number_format = 'R$ #,##0.00;[RED]-R$ #,##0.00'
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12  # codigo
    ws.column_dimensions['B'].width = 25  # cliente
    ws.column_dimensions['C'].width = 18  # valor
    ws.column_dimensions['D'].width = 15  # TIPO
    
    # Salvar as alterações
    wb.save(caminho_saida)
    print(f"\n✓ Planilha normalizada salva em: {caminho_saida}")


# ===================================================================
# EXEMPLOS DE USO
# ===================================================================

# OPÇÃO 1: Passar o caminho do arquivo
def exemplo_com_caminho():
    """Exemplo usando caminho do arquivo"""
    arquivo_entrada = "/mnt/user-data/uploads/financeiro.xlsx"
    
    # Normalizar passando o caminho
    df_normalizado = normalizar_planilha_financeira(arquivo_entrada)
    
    print(f"\nTotal de registros: {len(df_normalizado)}")
    return df_normalizado


# OPÇÃO 2: Passar o DataFrame já lido (SEU CASO)
def exemplo_com_dataframe():
    """Exemplo usando DataFrame já lido - RECOMENDADO PARA SEU CASO"""
    
    # Você faz a leitura antes (com sua função ler_arquivo)
    df_origem = pd.read_excel("/mnt/user-data/uploads/financeiro.xlsx")
    
    # Passa o DataFrame para a função
    df_normalizado = normalizar_planilha_financeira(df_origem)
    
    print(f"\nTotal de registros: {len(df_normalizado)}")
    return df_normalizado


# OPÇÃO 3: Seu código exato
def seu_exemplo():
    """Como ficaria no seu código"""
    
    def ler_arquivo(caminho):
        """Sua função de leitura"""
        return pd.read_excel(caminho)
    
    # Seu código
    origem_path = "/mnt/user-data/uploads/financeiro.xlsx"
    df_origem = ler_arquivo(origem_path)  # Você lê antes
    
    # Passa o DataFrame lido para a função
    financeiro_nor = normalizar_planilha_financeira(df_origem)  # ✅ FUNCIONA!
    
    return financeiro_nor


# Exemplo de execução
if __name__ == "__main__":
    print("="*70)
    print("TESTE: USANDO DATAFRAME JÁ LIDO")
    print("="*70)
    
    try:
        # Simular seu caso de uso
        origem_path = "/mnt/user-data/uploads/financeiro.xlsx"
        
        # Você lê o arquivo antes
        df_origem = pd.read_excel(origem_path)
        print(f"✓ Arquivo lido: {len(df_origem)} registros")
        
        # Passa o DataFrame para normalizar
        financeiro_nor = normalizar_planilha_financeira(df_origem)
        print(f"✓ Normalização concluída: {len(financeiro_nor)} registros agrupados")
        
        print("\n" + "="*70)
        print("PRIMEIRAS 5 LINHAS")
        print("="*70)
        print(financeiro_nor.head())
        
        # Salvar resultado
        salvar_planilha_normalizada(
            financeiro_nor, 
            "/mnt/user-data/outputs/financeiro_normalizado.xlsx"
        )
        
        print("\n✅ SUCESSO! A função aceita DataFrame já lido.")
        
    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        import traceback
        traceback.print_exc()
