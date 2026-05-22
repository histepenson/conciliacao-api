import io
from datetime import date
from decimal import Decimal
from sqlalchemy.orm import Session

from models.estoque import EstoqueSaldo, EstoqueMovimentacao, TipoMovimentacao
from models.produto import Produto
from services.estoque_service import listar_saldos, listar_movimentacoes, _primeiro_dia_mes


# ── Excel ────────────────────────────────────────────────────────────────────

def _estilo_header(ws, row: int, cols: int, fill_color: str = "1F4E79"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def gerar_excel_saldos(db: Session, empresa_id: int, periodo: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Saldos"

    periodo = _primeiro_dia_mes(periodo)

    ws.append([f"Relatório de Saldo de Estoque — {periodo.strftime('%m/%Y')}"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["Código", "Descrição", "Unidade", "Saldo Inicial", "Entradas", "Saídas", "Ajustes", "Saldo Final"]
    ws.append(headers)
    _estilo_header(ws, 3, len(headers))

    saldos = listar_saldos(db, empresa_id, periodo)
    for s in saldos:
        ws.append([
            getattr(s, "produto_codigo", ""),
            getattr(s, "produto_descricao", ""),
            getattr(s, "produto_unidade", ""),
            float(s.saldo_inicial),
            float(s.entradas),
            float(s.saidas),
            float(s.ajustes),
            float(s.saldo_final),
        ])
        row_idx = ws.max_row
        # Destaque saldo negativo
        if s.saldo_final < 0:
            neg_fill = PatternFill("solid", fgColor="FFCCCC")
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = neg_fill

    # Larguras
    col_widths = [14, 40, 10, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Formatar números
    for row in ws.iter_rows(min_row=4, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = '#,##0.0000'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_excel_movimentacoes(
    db: Session,
    empresa_id: int,
    produto_id: int | None = None,
    periodo: date | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    titulo = "Relatório de Movimentações"
    if periodo:
        titulo += f" — {periodo.strftime('%m/%Y')}"
    ws.append([titulo])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["Data", "Código", "Descrição", "Tipo", "Quantidade", "Documento", "Justificativa"]
    ws.append(headers)
    _estilo_header(ws, 3, len(headers))

    movs = listar_movimentacoes(db, empresa_id, produto_id, periodo)
    for m in movs:
        ws.append([
            m.data_movimentacao.strftime("%d/%m/%Y"),
            getattr(m, "produto_codigo", ""),
            getattr(m, "produto_descricao", ""),
            m.tipo.value,
            float(m.quantidade),
            m.documento_origem or "",
            m.justificativa or "",
        ])

    col_widths = [12, 14, 40, 12, 16, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=4, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.0000'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def gerar_pdf_saldos(db: Session, empresa_id: int, periodo: date) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    periodo = _primeiro_dia_mes(periodo)
    saldos = listar_saldos(db, empresa_id, periodo)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    story = [
        Paragraph(f"Relatório de Saldo de Estoque", title_style),
        Paragraph(f"Período: {periodo.strftime('%m/%Y')}", sub_style),
        Spacer(1, 8*mm),
    ]

    header = ["Código", "Descrição", "Unidade", "Saldo Inicial", "Entradas", "Saídas", "Ajustes", "Saldo Final"]
    data = [header]
    for s in saldos:
        def fmt(v): return f"{float(v):,.4f}"
        data.append([
            getattr(s, "produto_codigo", ""),
            getattr(s, "produto_descricao", ""),
            getattr(s, "produto_unidade", ""),
            fmt(s.saldo_inicial),
            fmt(s.entradas),
            fmt(s.saidas),
            fmt(s.ajustes),
            fmt(s.saldo_final),
        ])

    col_widths = [30*mm, 85*mm, 20*mm, 30*mm, 28*mm, 28*mm, 28*mm, 30*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (2, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FF")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])

    # Destaque vermelho em saldo negativo
    for i, s in enumerate(saldos, 1):
        if s.saldo_final < 0:
            style.add("BACKGROUND", (7, i), (7, i), colors.HexColor("#FFCCCC"))
            style.add("TEXTCOLOR", (7, i), (7, i), colors.HexColor("#CC0000"))

    t.setStyle(style)
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def gerar_pdf_movimentacoes(
    db: Session,
    empresa_id: int,
    produto_id: int | None = None,
    periodo: date | None = None,
) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    movs = listar_movimentacoes(db, empresa_id, produto_id, periodo)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    titulo = "Relatório de Movimentações de Estoque"
    subtitulo = f"Período: {periodo.strftime('%m/%Y')}" if periodo else "Todas as movimentações"

    story = [
        Paragraph(titulo, title_style),
        Paragraph(subtitulo, sub_style),
        Spacer(1, 8*mm),
    ]

    header = ["Data", "Código", "Descrição", "Tipo", "Quantidade", "Documento", "Justificativa"]
    data = [header]
    for m in movs:
        data.append([
            m.data_movimentacao.strftime("%d/%m/%Y"),
            getattr(m, "produto_codigo", ""),
            getattr(m, "produto_descricao", ""),
            m.tipo.value,
            f"{float(m.quantidade):,.4f}",
            m.documento_origem or "",
            (m.justificativa or "")[:60],
        ])

    col_widths = [22*mm, 25*mm, 75*mm, 22*mm, 25*mm, 35*mm, 75*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FF")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    story.append(t)
    doc.build(story)
    return buf.getvalue()
